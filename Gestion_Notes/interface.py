import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from gestion_etudiants import *
import openpyxl
from tkinter import filedialog
import os
dossier = os.path.dirname(os.path.abspath(__file__))
"""
from gestion_etudiants import *
from gestion_modules import *
from gestion_notes import *"""
NOTES_MODULES = {}
notes = []
#index des modules (note,coef)
index_modules = {"Algèbre 1":(4,5), "Analyse 1":(6,7), "MTU":(8,9),"P. Python":(10,11),"Algorithmique et p. C":(12,13),"Architecture fonctionnement des ordinateurs":(14,15),"Électronique numérique":(16,17)}
# not(entrer des notes prenierment)
def verifier_av():
    if len(ETUDIANTS) > 0 and len(COEFFICIENTS) == 7 :
        bouton_note.config(state="normal")
    else:
        bouton_note.config(state="disabled")
#import gestion_etudiants as etudiants
#import gestion_modules as modules
fenetre = tk.Tk()
fenetre.title("Application de Gestion de Notes.")
fenetre.geometry("1690x780+1+1")
fenetre.configure(background="#E5E4E2")
#titre
Title = tk.Label(fenetre,text="Gestion de Notes",bg="dark blue",font=("monospace",17,"bold"),fg="white")
Title.pack(fill="x")
#frame menu
part_frame = tk.Frame(fenetre,bg="white",width=300,height=400)
part_frame.pack(side="left",fill="y")
part_frame.pack_propagate(False)
#entre nom ....
vcmd_nom_prenom = fenetre.register(valider_nom_prenom)
vcmd_cne = fenetre.register(valider_cne)
def Boutton_ajoute():
    NOM = Entrer_nom.get().strip()
    PRENOM = Entrer_prenom.get().strip()
    CNE = Entrer_cne.get().strip()
    if NOM == "" or PRENOM == "" or CNE == "" :
        messagebox.showinfo("Erreur","Remplir les cas.")
        return
    etudiant = ajouter_etudiant(NOM,PRENOM,CNE)
    if etudiant is None :
        messagebox.showinfo("Erreur","CNE deja existant!")
        return
    donnee_table.insert("","end",values=(etudiant["Id"],etudiant["Nom"],etudiant["Prenom"],etudiant["cne"],"","","","","","","","","","","","","","",0,0,""))
    combo_box_etd["values"] = get_noms_complets()
    Entrer_nom.delete(0,tk.END)
    Entrer_prenom.delete(0,tk.END)
    Entrer_cne.delete(0,tk.END)
    verifier_av()
def Boutton_supprimer():
    selection = donnee_table.selection()
    if not selection:
        return
    values = donnee_table.item(selection)["values"]
    id_supp = values[0]
    supprimer_etudiant(id_supp)
    donnee_table.delete(selection)
    combo_box_etd["values"] = get_noms_complets()
    combo_box_etd.set("")
    
def Boutton_modifier():
    selection = donnee_table.selection()
    if not selection:
        messagebox.showinfo("Erreur","Selectionner un etudiant!")
        return
    NOM = Entrer_nom.get().strip()
    PRENOM = Entrer_prenom.get().strip()
    CNE = Entrer_cne.get().strip()
    if NOM == "" or PRENOM == "" or CNE == "" :
        messagebox.showinfo("Erreur","Remplir les cas.")
        return
    values = donnee_table.item(selection)["values"]
    id_mod = values[0]
    for etudiant in ETUDIANTS :
        if str(etudiant["Id"]) == str(id_mod):
            etudiant["Nom"] = NOM
            etudiant["Prenom"] = PRENOM
            etudiant["cne"] = CNE
            break
    old_values = list(donnee_table.item(selection)["values"])
    old_values[1] = NOM
    old_values[2] = PRENOM
    old_values[3] = CNE
    donnee_table.item(selection,values=old_values)
    liste_nom_complet.clear()
    for e in ETUDIANTS :
        liste_nom_complet.append(e["Nom"] + " " + e["Prenom"])
    combo_box_etd["values"] = liste_nom_complet
    combo_box_etd.set("")
    Entrer_nom.delete(0,tk.END)
    Entrer_prenom.delete(0,tk.END)
    Entrer_cne.delete(0,tk.END)
LBL_nom = tk.Label(part_frame,text="Etudians :",bg="white",font=("helvetica",10,"bold"))
LBL_nom.place(x=1,y=1)
LBL_nom = tk.Label(part_frame,text="Nom :",bg="white")
LBL_nom.place(x=55,y=20)
Entrer_nom = tk.Entry(part_frame,bd="2",justify="center",validate="key",validatecommand=(vcmd_nom_prenom,"%P"))
Entrer_nom.place(x=95,y=20)
LBL_prenom = tk.Label(part_frame,text="Prenom :",bg="white")
LBL_prenom.place(x=40,y=45)
Entrer_prenom = tk.Entry(part_frame,bd="2",justify="center",validate="key",validatecommand=(vcmd_nom_prenom,"%P"))
Entrer_prenom.place(x=95,y=45)
LBL_cne = tk.Label(part_frame,text="CNE :",bg="white")
LBL_cne.place(x=59,y=70)
Entrer_cne = tk.Entry(part_frame,bd="2",justify="center",validate="key",validatecommand=(vcmd_cne,"%P"))
Entrer_cne.place(x=95,y=70)
#botton des ajoute et supp et modifier
btn_ajoute = tk.Button(part_frame,text="Ajouter",bg="#3EB489",command=Boutton_ajoute)
btn_ajoute.place(x=125,y=95)
LBL_selec = tk.Label(part_frame,text="Selectioner :",bg="white")
LBL_selec.place(x=1,y=126)
btn_supp = tk.Button(part_frame,text="Supprimer",bg="#CF1020",command=Boutton_supprimer)
btn_supp.place(x=80,y=125)
btn_mod = tk.Button(part_frame,text="Modifier",bg="#0FC0FC",command=Boutton_modifier)
btn_mod.place(x=152,y=125)
canvas_1 = tk.Canvas(part_frame,height=2)
canvas_1.place(x=1,y=155)
canvas_1.create_line(0,1,400,1)
#base de donnee 
colonnes = ("Id","Nom","Prenom","CNE","Algèbre 1","Coeff_Alg","Analyse 1","Coeff_Ana","MTU","Coeff_MTU","P. Python","Coeff_Py","Algo et p. C","Coeff_Algo","Archi des ordinateurs","Coeff_Arc","Électronique num","Coeff_E_N","Total","Moyenne","Montion")
donnees_frame = tk.Frame(fenetre,bg="#F5F5F5",width=1250,height=600)
donnees_frame.pack(side="left",fill="both",expand=True)
scroll_donnees_x = tk.Scrollbar(donnees_frame,orient=tk.HORIZONTAL)
scroll_donnees_y = tk.Scrollbar(donnees_frame,orient=tk.VERTICAL)
donnee_table = ttk.Treeview(donnees_frame,columns= colonnes,show="headings",height=10,xscrollcommand=scroll_donnees_x.set,yscrollcommand=scroll_donnees_y.set)
scroll_donnees_x.pack(side="bottom",fill="x")
scroll_donnees_y.pack(side="right",fill="y")
donnee_table.pack(fill="both",expand=True)
scroll_donnees_x.config(command=donnee_table.xview)
scroll_donnees_y.config(command=donnee_table.yview)
#titre de tableau de base donner
donnee_table.heading("Id",text="ID")
donnee_table.heading("Nom",text="Nom")
donnee_table.heading("Prenom",text="Prenom")
donnee_table.heading("CNE",text="CNE")
donnee_table.heading("Algèbre 1",text="Algèbre 1")
donnee_table.heading("Coeff_Alg",text="Coeff_Alg")
donnee_table.heading("Analyse 1",text="Analyse 1")
donnee_table.heading("Coeff_Ana",text="Coeff_Ana")
donnee_table.heading("MTU",text="MTU")
donnee_table.heading("Coeff_MTU",text="Coeff_MTU")
donnee_table.heading("P. Python",text="P. Python")
donnee_table.heading("Coeff_Py",text="Coeff_Py")
donnee_table.heading("Algo et p. C",text="Algo et p. C")
donnee_table.heading("Coeff_Algo",text="Coeff_Algo")
donnee_table.heading("Archi des ordinateurs",text="Archi des ordinateurs")
donnee_table.heading("Coeff_Arc",text="Coeff_Arc")
donnee_table.heading("Électronique num",text="Électronique num")
donnee_table.heading("Coeff_E_N",text="Coeff_E_N")
donnee_table.heading("Total",text="Total")
donnee_table.heading("Moyenne",text="Moyenne")
donnee_table.heading("Montion",text="Montion")

donnee_table.column("Id",width=30,anchor="center")
donnee_table.column("Nom",width=50,anchor="center")
donnee_table.column("Prenom",width=60,anchor="center")
donnee_table.column("CNE",width=70,anchor="center")
donnee_table.column("Algèbre 1",width=70,anchor="center")
donnee_table.column("Coeff_Alg",width=80,anchor="center")
donnee_table.column("Analyse 1",width=70,anchor="center")
donnee_table.column("Coeff_Ana",width=70,anchor="center")
donnee_table.column("MTU",width=40,anchor="center")
donnee_table.column("Coeff_MTU",width=70,anchor="center")
donnee_table.column("P. Python",width=70,anchor="center")
donnee_table.column("Coeff_Py",width=70,anchor="center")
donnee_table.column("Algo et p. C",width=70,anchor="center")
donnee_table.column("Coeff_Algo",width=70,anchor="center")
donnee_table.column("Archi des ordinateurs",width=130,anchor="center")
donnee_table.column("Coeff_Arc",width=70,anchor="center")
donnee_table.column("Électronique num",width=120,anchor="center")
donnee_table.column("Coeff_E_N",width=70,anchor="center")
donnee_table.column("Total",width=70,anchor="center")
donnee_table.column("Moyenne",width=80,anchor="center")
donnee_table.column("Montion",width=80,anchor="center")
# entrer les coeff
COEFFICIENTS = {}
LBL_Coeff_mdl = tk.Label(part_frame,text="Coefficient de modules :",bg="white",font=("helvetica",10,"bold"))
LBL_Coeff_mdl.place(x=1,y=160)
LBL_mdl = tk.Label(part_frame, text="Les modules :",bg="white")
LBL_mdl.place(x=120,y=190)
combo_box = ttk.Combobox(part_frame,values=["Algèbre 1", "Analyse 1", "MTU","P. Python","Algorithmique et p. C","Architecture fonctionnement des ordinateurs","Électronique numérique"],state="readonly",width=35,)
combo_box.set("Algèbre 1")
combo_box.place(x=40,y=210)
LBL_coeff = tk.Label(part_frame, text="Coefficient : ",bg="white")
LBL_coeff.place(x=120,y=240)
spinbox = tk.Spinbox(part_frame,from_=1, to= 10,bd=2)
spinbox.place(x=80,y=260)
table_coeff = ttk.Treeview(part_frame, columns=("Module", "Coefficient"), show="headings",height=7)
table_coeff.heading("Module", text="Module")
table_coeff.heading("Coefficient", text="Coefficient")
table_coeff.column("Module", width=100)
table_coeff.column("Coefficient", width=100, anchor="center")
table_coeff.place(x=40,y=295)
#complet table_coeff menu
def ajouter_Coeff():
    if len(COEFFICIENTS) >= 7 :
        messagebox.showinfo("Information", "Tous les Coefficients sont entrés.")
        btn.config(state="disabled")
        return
    module = combo_box.get()
    coefficient = int(spinbox.get())
    COEFFICIENTS[module] = coefficient
    table_coeff.insert("", "end", values=(module, coefficient))
    current_values = list(combo_box["values"])
    if module in current_values:
        current_values.remove(module)
        combo_box["values"] = current_values
    if current_values:
        combo_box.set(current_values[0])
    else:
        combo_box.set("")
    verifier_av()
bouton_coff = tk.Button(part_frame,text="sauvegarder",activebackground="white",bg="#3EB489",command=ajouter_Coeff)
bouton_coff.place(x=100,y=461)
canvas_2 = tk.Canvas(part_frame,height=2)
canvas_2.place(x=1,y=498)
canvas_2.create_line(0,1,400,1)



#Entrer la note de chaque module
        #verification de note
def verifier_note(text_actuuel):
    if text_actuuel == "":
        return True
    if text_actuuel == ".":
        return True
    if text_actuuel.count(".") > 1 :
        return False
    if "." in text_actuuel :
        apres_point = text_actuuel.split(".")[1]
        if len(apres_point) > 2:
            return False
    try:
        nbr = float(text_actuuel)
        if nbr > 20:
            return False
        if nbr < 0 :
            return False
        return True
    except:
        return False
          #stocker la note 
def on_select_etudiant(event):
    etudiant_nom = combo_box_etd.get()
    etudiant = None
    for e in ETUDIANTS:
        if e["Nom"] + " " + e["Prenom"] == etudiant_nom:
            etudiant = e
            break
    for item in donnee_table.get_children():
        valeurs = list(donnee_table.item(item)["values"])
        if str(valeurs[0]) == str(etudiant["Id"]):
            modules_restants = []
            for module in index_modules:
                idx_n, idx_c = index_modules[module]
                if str(valeurs[idx_n]) == "" or str(valeurs[idx_n]) == "0":
                    modules_restants.append(module)
            combo_box_note["values"] = modules_restants
            if modules_restants:
                combo_box_note.set(modules_restants[0])
            else:
                combo_box_note.set("")
            break
def stocker_notes():
    etudiant_nom = combo_box_etd.get()
    module = combo_box_note.get()
    note = Entrer_note.get()
    if combo_box_note.get() == "" :
        messagebox.showinfo("Info", "Toutes les Notes sont saisies pour cet etudiant!")
        Entrer_note.delete(0,tk.END)
        return
    if  etudiant_nom == "" :
        messagebox.showinfo("Erreur","Choisir l'etudiant!")
        return
    if module == "" or note == "" or etudiant_nom == "" :
        messagebox.showinfo("Erreur","Remplir les notes de chaque modules")
        return
    coef = COEFFICIENTS[module]
    note = float(note)
    etudiant = None
    for e in ETUDIANTS:
        if e["Nom"] + " " + e["Prenom"] == etudiant_nom:
            etudiant = e
            break
    for item in donnee_table.get_children():
        valeurs = list(donnee_table.item(item)["values"])
        if str(valeurs[0]) == str(etudiant["Id"]):
            index_note , index_coef = index_modules[module]
            valeurs[index_note] = note
            valeurs[index_coef] = coef
            donnee_table.item(item,values=valeurs)
            break
    current_values = list(combo_box_note["values"])
    if module in current_values:
        current_values.remove(module)
        combo_box_note["values"] = current_values
    if current_values:
        combo_box_note.set(current_values[0])
    else:
        combo_box_note.set("")                    
    Entrer_note.delete(0,tk.END)
    Tous_remplis = True
    for mdl in index_modules:
        idx_n , idx_c = index_modules[mdl]
        if str(valeurs[idx_n]) == "" :
            Tous_remplis = False
            break
        if Tous_remplis:
            Total = 0
            Total_coef = 0
            for m in index_modules:
                i_n , i_c = index_modules[m]
                Total += float(valeurs[i_n]) * int(valeurs[i_c]) 
                Total_coef += int(valeurs[i_c])
            Moyenne = float(format(Total / Total_coef,".2f"))
            valeurs[18] = float(format(Total,".2f"))
            valeurs[19] = Moyenne
            if Moyenne < 10 :
                valeurs[20] = "N.valide"
            elif Moyenne < 12 :
                valeurs[20] = "Admis"
            elif Moyenne < 14 :
                valeurs[20] = "A.bien"
            elif Moyenne < 16 :
                valeurs[20] = "Bien"
            else:
                valeurs[20] = "T.bien"
            donnee_table.item(item,values=valeurs)


LBL_note_mdl = tk.Label(part_frame,text="Note du module :",bg="white",font=("helvetica",10,"bold"))
LBL_note_mdl.place(x=1,y=505)
LBL_choi_etd = tk.Label(part_frame, text="Etudiants :",bg="white")
LBL_choi_etd.place(x=15,y=535)
combo_box_etd = ttk.Combobox(part_frame,values=ETUDIANTS,state="readonly",width=30)
combo_box_etd.set("")
combo_box_etd.bind("<<ComboboxSelected>>", on_select_etudiant)
combo_box_etd.place(x=75,y=535)
LBL_choi_mdl = tk.Label(part_frame, text="Les modules :",bg="white")
LBL_choi_mdl.place(x=1,y=565)
combo_box_note = ttk.Combobox(part_frame,values=["Algèbre 1", "Analyse 1", "MTU","P. Python","Algorithmique et p. C","Architecture fonctionnement des ordinateurs","Électronique numérique"],state="readonly",width=30)
combo_box_note.set("Algèbre 1")
combo_box_note.place(x=78,y=565)
LBL_ent_note = tk.Label(part_frame,text="Note :",bg="white")
LBL_ent_note.place(x=100,y=600)
verf = part_frame.register(verifier_note)
Entrer_note = tk.Entry(part_frame,bd="2",justify="center",width=8,validate="key",validatecommand=(verf,"%P"))
Entrer_note.place(x=140,y=600)
bouton_note = tk.Button(part_frame,text="sauvegarder",activebackground="white",bg="#3EB489",command=stocker_notes,state="disabled")
bouton_note.place(x=127,y=640)
canvas_3 = tk.Canvas(part_frame,height=2)
canvas_3.place(x=1,y=670)
canvas_3.create_line(0,1,400,1)
LBL_excel = tk.Label(part_frame,text="Excel :",bg="white",font=("helvetica",10,"bold"))
LBL_excel.place(x=1,y=680)
LBL_releve = tk.Label(part_frame,text="Relevé de Notes :",bg="white",font=("helvetica",10,"bold"))
LBL_releve.place(x=1,y=715)
def ouvrir_releve():
    top = tk.Toplevel(fenetre)
    top.title("Relevé de Notes")
    top.geometry("400x400")
    top.configure(bg="white")
    tk.Label(top, text="Choisir un étudiant:",bg="white", font=("helvetica", 12, "bold")).pack(pady=10)
    # Listbox
    listbox = tk.Listbox(top, width=40, height=10, font=("helvetica", 11))
    listbox.pack(pady=10)
    etudiants_complets = []
    for e in ETUDIANTS:
        for item in donnee_table.get_children():
            valeurs = list(donnee_table.item(item)["values"])
            if str(valeurs[0]) == str(e["Id"]) :
                complet = True
                for module in index_modules:
                    idx_n , idx_c = index_modules[module]
                    print("m : ",module,"notes :",valeurs[idx_n],type(valeurs[idx_n]))
                    try:
                        float(valeurs[idx_n])
                    except:
                        print("PAS UN NOMBRE")
                        complet = False
                        break
                if complet:
                    etudiants_complets.append(e)
                    listbox.insert("end",e["Nom"] + " " + e["Prenom"] + " - " + str(e["cne"]))
                    break
                print("COMPLET : ",complet)
    if len(etudiants_complets) == 0 :
        messagebox.showinfo("Info","Aucun etudiant n'a toutes ses notes!")
        top.destroy()
        return
    
    def generer_releve():
        selection = listbox.curselection()
        if not selection:
            messagebox.showinfo("Erreur", "Choisir un étudiant!")
            return
        index_sel = selection[0]
        etudiant = ETUDIANTS[index_sel]
        for item in donnee_table.get_children():
            valeurs = list(donnee_table.item(item)["values"])
            if str(valeurs[0]) == str(etudiant["Id"]):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Relevé de Notes"
                ws.merge_cells("A1:D1")
                ws["A1"] = "Relevé de Notes"
                ws["A1"].font = openpyxl.styles.Font(size=16, bold=True)
                
                # info_etudiant
                ws["A3"] = "Nom:"
                ws["B3"] = etudiant["Nom"]
                ws["A4"] = "Prénom:"
                ws["B4"] = etudiant["Prenom"]
                ws["A5"] = "CNE:"
                ws["B5"] = etudiant["cne"]
                ws["A7"] = "Module"
                ws["B7"] = "Note"
                ws["C7"] = "Coefficient"
                ws["A7"].font = openpyxl.styles.Font(bold=True)
                ws["B7"].font = openpyxl.styles.Font(bold=True)
                ws["C7"].font = openpyxl.styles.Font(bold=True)
    
                ligne = 8
                for module in index_modules:
                    idx_n, idx_c = index_modules[module]
                    ws[f"A{ligne}"] = module
                    ws[f"B{ligne}"] = valeurs[idx_n]
                    ws[f"C{ligne}"] = valeurs[idx_c]
                    ligne += 1 
                # Total, Moyenne et Mention
                ligne += 1
                ws[f"A{ligne}"] = "Total:"
                ws[f"B{ligne}"] = valeurs[18]
                ligne += 1
                ws[f"A{ligne}"] = "Moyenne:"
                ws[f"B{ligne}"] = valeurs[19]
                ligne += 1
                ws[f"A{ligne}"] = "Mention:"
                ws[f"B{ligne}"] = valeurs[20]
                ws.column_dimensions["A"].width = 40
                ws.column_dimensions["B"].width = 15
                ws.column_dimensions["C"].width = 15
                nom_fichier = f"Releve_{etudiant['Nom']}_{etudiant['Prenom']}.xlsx"
                chemin = os.path.join(dossier,nom_fichier)
                wb.save(chemin)
                messagebox.showinfo("Succès", f"Relevé sauvegardé: {nom_fichier}")
                top.destroy()
                break
    btn_generer = tk.Button(top, text="Générer Relevé",bg="#1976d2", fg="white",font=("helvetica", 11, "bold"),command=generer_releve)
    btn_generer.pack(pady=20)
btn_releve = tk.Button(part_frame, text="Relevé",bg="#e65100", fg="white",command=ouvrir_releve)
btn_releve.place(x=135, y=715)
def btn_sauvegarder():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Étudiants"
    
    titres = ["ID", "Nom", "Prenom", "CNE","Algèbre 1", "Coeff","Analyse 1", "Coeff","MTU", "Coeff","P. Python", "Coeff","Algo et p. C", "Coeff","Archi des ordinateurs", "Coeff","Électronique num", "Coeff","Total", "Moyenne", "Mention"]
    for col_num, titre in enumerate(titres, 1):
        ws.cell(row=1, column=col_num).value = titre
    row_num = 2
    for item in donnee_table.get_children():
        valeurs = list(donnee_table.item(item)["values"])
        for col_num, v in enumerate(valeurs, 1):
            v_str = str(v).strip()
            if v_str != "":
                ws.cell(row=row_num, column=col_num).value = v_str
        row_num += 1
    ws2 = wb.create_sheet("Coefficients")
    ws2.append(["Module", "Coefficient"])
    for module, coef in COEFFICIENTS.items():
        ws2.append([module, coef])
    chemin = os.path.join(dossier, "donnees_notes.xlsx")
    wb.save(chemin)
    messagebox.showinfo("Succès", "Données sauvegardées!")
def btn_charger():
    fichier = filedialog.askopenfilename(title="Choisir un fichier",filetypes=[("Excel", "*.xlsx")])
    if fichier == "":
        return
    wb = openpyxl.load_workbook(fichier)
    for item in donnee_table.get_children():
        donnee_table.delete(item)
    ETUDIANTS.clear()
    COEFFICIENTS.clear()
    liste_nom_complet.clear()
    for item in table_coeff.get_children():
        table_coeff.delete(item)
    ws1 = wb["Étudiants"]
    for i, row in enumerate(ws1.iter_rows(values_only=True)):
        if i == 0:
            continue
        donnee_table.insert("", "end", values=list(row))
        etudiant = {"Id": row[0],"Nom": row[1],"Prenom": row[2],"cne": row[3]}
        ETUDIANTS.append(etudiant)
    ws2 = wb["Coefficients"]
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i == 0:
            continue
        COEFFICIENTS[row[0]] = row[1]
        table_coeff.insert("", "end", values=(row[0], row[1]))
    for e in ETUDIANTS:
        liste_nom_complet.append(e["Nom"] + " " + e["Prenom"])
    combo_box_etd["values"] = liste_nom_complet
    restants = []
    for m in ["Algèbre 1", "Analyse 1", "MTU", "P. Python","Algorithmique et p. C","Architecture fonctionnement des ordinateurs","Électronique numérique"]:
        if m not in COEFFICIENTS:
            restants.append(m)
    combo_box["values"] = restants
    if restants:
        combo_box.set(restants[0])
    else:
        combo_box.set("")
    verifier_av()
    messagebox.showinfo("Succès", "Données chargées!")
btn_save = tk.Button(part_frame, text="Sauvegarder", bg="#2e7d32",fg="white",command=btn_sauvegarder)
btn_save.place(x=70, y=680)
btn_charger = tk.Button(part_frame, text="Charger", bg="#1565c0",fg="white",command=btn_charger)
btn_charger.place(x=160, y=680)
# Auto de charger les donnees
chemin = os.path.join(dossier,"donnees_notes.xlsx")
if os.path.exists(chemin) :
    wb = openpyxl.load_workbook(chemin)
    ws1 = wb["Étudiants"]
    for i, row in enumerate(ws1.iter_rows(values_only=True)):
        if i == 0:
            continue
        ligne = []
        for v in row :
            if v is None :
                ligne.append("")
            else:
                ligne.append(v)
        donnee_table.insert("", "end", values=ligne)
        etudiant = {"Id": row[0],"Nom": row[1],"Prenom": row[2],"cne": row[3]}
        ETUDIANTS.append(etudiant)
    ws2 = wb["Coefficients"]
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i == 0:
            continue
        COEFFICIENTS[row[0]] = row[1]
        table_coeff.insert("", "end", values=(row[0], row[1]))
    liste_nom_complet.clear()
    for e in ETUDIANTS:
        liste_nom_complet.append(e["Nom"] + " " + e["Prenom"])
    combo_box_etd["values"] = liste_nom_complet
    restants = []
    for m in ["Algèbre 1", "Analyse 1", "MTU", "P. Python","Algorithmique et p. C","Architecture fonctionnement des ordinateurs","Électronique numérique"]:
        if m not in COEFFICIENTS:
            restants.append(m)
    combo_box["values"] = restants
    if restants:
        combo_box.set(restants[0])
    else:
        combo_box.set("")
    combo_box_note["values"] = ["Algèbre 1", "Analyse 1", "MTU", "P. Python","Algorithmique et p. C","Architecture fonctionnement des ordinateurs","Électronique numérique"]
    combo_box_note.set("")
    if len(ETUDIANTS) > 0 :
        max_id = 0
        for e in ETUDIANTS :
            if int(e["Id"]) > max_id :
                max_id = int(e["Id"])
        ID_auto = max_id + 1 
    verifier_av()
fenetre.mainloop()