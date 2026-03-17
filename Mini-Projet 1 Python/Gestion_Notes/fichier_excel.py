import openpyxl
import os

def sauvegarder_excel(donnee_table, coefficients, dossier):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Étudiants"
    titres = ["ID", "Nom", "Prenom", "CNE","Algèbre 1", "Coef","Analyse 1", "Coef","MTU", "Coef","P. Python", "Coef","Algo et p. C", "Coef","Archi. des ordinateurs", "Coef","Électronique num", "Coef","Total", "Moyenne", "Mention"]
    for col_num, titre in enumerate(titres, 1):
        ws.cell(row=1, column=col_num).value = titre
    row_num = 2
    for item in donnee_table.get_children():
        valeurs = list(donnee_table.item(item)["values"])
        for col_num, v in enumerate(valeurs, 1):
            v_str = str(v).strip()
            if v_str != "":
                ws.cell(row=row_num, column=col_num).value = v_str
        row_num += 1
    ws2 = wb.create_sheet("Coefficients")
    ws2.append(["Module", "Coefficient"])
    for module, coef in coefficients.items():
        ws2.append([module, coef])
    chemin = os.path.join(dossier, "donnees_notes.xlsx")
    wb.save(chemin)
    return True

def charger_excel(fichier):
    wb = openpyxl.load_workbook(fichier)
    ws1 = wb["Étudiants"]
    etudiants = []
    lignes = []
    for i, row in enumerate(ws1.iter_rows(values_only=True)):
        if i == 0:
            continue
        ligne = []
        for v in row:
            if v is None:
                ligne.append("")
            else:
                ligne.append(v)
        lignes.append(ligne)
        etudiant = {"Id": row[0],"Nom": row[1],"Prenom": row[2],"cne": row[3]}
        etudiants.append(etudiant)
    ws2 = wb["Coefficients"]
    coefficients = {}
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i == 0:
            continue
        coefficients[row[0]] = row[1]
    return etudiants, coefficients, lignes

def generer_releve(etudiant, valeurs, index_modules, dossier):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relevé de Notes"
    ws.merge_cells("A1:D1")
    ws["A1"] = "Relevé de Notes"
    ws["A1"].font = openpyxl.styles.Font(size=16, bold=True)
    ws["A3"] = "Nom:"
    ws["B3"] = etudiant["Nom"]
    ws["A4"] = "Prénom:"
    ws["B4"] = etudiant["Prenom"]
    ws["A5"] = "CNE:"
    ws["B5"] = str(etudiant["cne"])
    ws["A7"] = "Module"
    ws["B7"] = "Note"
    ws["C7"] = "Coefficient"
    ws["A7"].font = openpyxl.styles.Font(bold=True)
    ws["B7"].font = openpyxl.styles.Font(bold=True)
    ws["C7"].font = openpyxl.styles.Font(bold=True)
    ligne = 8
    for module in index_modules:
        idx_n, idx_c = index_modules[module]
        ws.cell(row=ligne, column=1).value = module
        ws.cell(row=ligne, column=2).value = valeurs[idx_n]
        ws.cell(row=ligne, column=3).value = valeurs[idx_c]
        ligne += 1
    ligne += 1
    ws.cell(row=ligne, column=1).value = "Total:"
    ws.cell(row=ligne, column=2).value = valeurs[18]
    ligne += 1
    ws.cell(row=ligne, column=1).value = "Moyenne:"
    ws.cell(row=ligne, column=2).value = valeurs[19]
    ligne += 1
    ws.cell(row=ligne, column=1).value = "Mention:"
    ws.cell(row=ligne, column=2).value = valeurs[20]
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    nom_fichier = "Releve_" + str(etudiant["Nom"]) + "_" + str(etudiant["Prenom"]) + ".xlsx"
    chemin = os.path.join(dossier, nom_fichier)
    wb.save(chemin)
    return nom_fichier